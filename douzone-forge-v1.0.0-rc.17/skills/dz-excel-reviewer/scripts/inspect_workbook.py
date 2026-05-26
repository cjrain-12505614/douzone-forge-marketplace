#!/usr/bin/env python3
"""
Inspect workbook structure with openpyxl first, then zip/xml fallback.

Usage:
    inspect_workbook.py <path-to-xlsx>
"""

from __future__ import annotations

import argparse
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


@dataclass
class SheetMeta:
    name: str
    state: str
    rid: str
    target: str | None = None


def inspect_with_openpyxl(path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    print(f"# Workbook Summary: {path.name}")
    print()
    print("- mode: openpyxl")
    print(f"- sheets: {len(wb.sheetnames)}")
    print(f"- sheet_names: {', '.join(wb.sheetnames)}")
    print()

    for ws in wb.worksheets:
        hidden_cols = sum(1 for d in ws.column_dimensions.values() if d.hidden)
        hidden_rows = sum(1 for d in ws.row_dimensions.values() if d.hidden)
        merged = len(list(ws.merged_cells.ranges))
        validations = len(getattr(ws.data_validations, "dataValidation", []))
        print(f"## {ws.title}")
        print(f"- state: {ws.sheet_state}")
        print(f"- max_row: {ws.max_row}")
        print(f"- max_col: {ws.max_column}")
        print(f"- hidden_rows: {hidden_rows}")
        print(f"- hidden_cols: {hidden_cols}")
        print(f"- merged_cells: {merged}")
        print(f"- data_validations: {validations}")
        print()
    return 0


def parse_workbook_meta(path: Path) -> list[SheetMeta]:
    with zipfile.ZipFile(path) as z:
        workbook = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))

        rel_map = {}
        for rel in rels:
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target")
            if rel_id and target:
                rel_map[rel_id] = target

        ns = {"a": MAIN_NS, "r": REL_NS}
        out: list[SheetMeta] = []
        for sheet in workbook.find("a:sheets", ns):
            rid = sheet.attrib.get(f"{{{REL_NS}}}id", "")
            out.append(
                SheetMeta(
                    name=sheet.attrib.get("name", ""),
                    state=sheet.attrib.get("state", "visible"),
                    rid=rid,
                    target=rel_map.get(rid),
                )
            )
        return out


def count_tag(data: bytes, tag: bytes) -> int:
    return data.count(tag)


def inspect_with_zipxml(path: Path) -> int:
    sheets = parse_workbook_meta(path)
    print(f"# Workbook Summary: {path.name}")
    print()
    print("- mode: zipxml-fallback")
    print(f"- sheets: {len(sheets)}")
    print(f"- visible_sheets: {sum(1 for s in sheets if s.state == 'visible')}")
    print(f"- hidden_sheets: {sum(1 for s in sheets if s.state != 'visible')}")
    print()

    with zipfile.ZipFile(path) as z:
        for sheet in sheets:
            print(f"## {sheet.name}")
            print(f"- state: {sheet.state}")
            print(f"- target: {sheet.target or 'unresolved'}")
            if not sheet.target:
                print()
                continue

            worksheet_path = "xl/" + sheet.target.lstrip("/")
            try:
                data = z.read(worksheet_path)
            except KeyError:
                print("- worksheet_xml: missing")
                print()
                continue

            print(f"- merged_cells: {count_tag(data, b'<mergeCell ')}")
            print(f"- data_validations: {count_tag(data, b'<dataValidation ')}")
            print(f"- hidden_rows: {count_tag(data, b' hidden=\"1\"')}")
            print()
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    args = parser.parse_args()
    path = Path(args.path)

    try:
        return inspect_with_openpyxl(path)
    except Exception as exc:  # noqa: BLE001
        print(f"# openpyxl failed: {exc}", file=sys.stderr)
        return inspect_with_zipxml(path)


if __name__ == "__main__":
    raise SystemExit(main())
